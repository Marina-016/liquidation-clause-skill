#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
清盘条款分类 — 三级递进流水线主入口。

阶段一: Datayes 基金合同 → PyMuPDF
阶段二: 替代公告源(招募说明书/发售公告等) → PyMuPDF+pypdf
阶段三: CSRC 证监会兜底 → pypdf

用法: python pipeline.py <基金列表.xlsx> [--output <输出.xlsx>] [--work-dir <工作目录>]
"""

import os, sys, json, time, argparse


CACHE_SCHEMA_VERSION = 2
PIPELINE_VERSION = "0.2.0"
CHECKPOINT_INTERVAL = 25


def save_results_cache(cache_path: str, results: list) -> None:
    """原子保存版本化结果缓存，避免中断时留下半个 JSON 文件。"""
    payload = {
        "schemaVersion": CACHE_SCHEMA_VERSION,
        "pipelineVersion": PIPELINE_VERSION,
        "savedAt": time.strftime("%Y-%m-%dT%H:%M:%S%z"),
        "results": results,
    }
    temp_path = f"{cache_path}.tmp"
    with open(temp_path, "w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)
        handle.flush()
        os.fsync(handle.fileno())
    os.replace(temp_path, cache_path)


def load_cached_results(cache_path: str, funds: list) -> tuple:
    """只复用当前版本缓存中已成功分类的基金，失败记录继续重试。"""
    if not os.path.exists(cache_path):
        return [], list(funds)

    try:
        with open(cache_path, "r", encoding="utf-8") as handle:
            payload = json.load(handle)
    except (OSError, ValueError):
        return [], list(funds)

    if not isinstance(payload, dict):
        return [], list(funds)
    if payload.get("schemaVersion") != CACHE_SCHEMA_VERSION:
        return [], list(funds)
    if payload.get("pipelineVersion") != PIPELINE_VERSION:
        return [], list(funds)

    cached_results = payload.get("results")
    if not isinstance(cached_results, list):
        return [], list(funds)

    cached_by_code = {
        str(result.get("code")): result
        for result in cached_results
        if isinstance(result, dict) and result.get("clauseType")
    }
    reused = []
    pending = []
    for fund in funds:
        code, name, mgr, type1, type2 = fund
        cached = cached_by_code.get(code)
        if not cached:
            pending.append(fund)
            continue
        refreshed = dict(cached)
        refreshed.update(
            {
                "code": code,
                "name": name,
                "mgr": mgr,
                "type1": type1,
                "type2": type2,
            }
        )
        reused.append(refreshed)
    return reused, pending


def filter_unclassified_funds(funds: list, *result_groups: list) -> list:
    """排除所有已有成功分类结果的代码，作为下一阶段的硬性入口过滤。"""
    classified_codes = {
        str(result.get("code"))
        for group in result_groups
        for result in group
        if isinstance(result, dict) and result.get("clauseType")
    }
    return [fund for fund in funds if fund[0] not in classified_codes]


def make_unclassified_result(fund: tuple) -> dict:
    code, name, mgr, type1, type2 = fund
    return {
        "code": code,
        "name": name,
        "mgr": mgr,
        "type1": type1,
        "type2": type2,
        "clauseType": None,
        "clauseText": "",
        "s3Url": "",
        "source": "",
        "stage": 3,
        "reason": "三阶段均未成功分类",
    }

def read_fund_list(xlsx_path: str) -> list:
    """读取基金列表 Excel，返回 [(code, name, mgr, type1, type2), ...]"""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    # 自动探测列名(模糊匹配)
    col_map = {}
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(row=1, column=c).value or "").strip()
        if not h:
            continue
        for key in ["代码", "基金代码", "fundCode", "code"]:
            if key.lower() in h.lower():
                col_map["code"] = c
        for key in ["名称", "基金名称", "简称", "fundName", "name"]:
            if key.lower() in h.lower():
                col_map["name"] = c
        for key in ["管理人", "基金管理人", "mgr", "manager"]:
            if key.lower() in h.lower():
                col_map["mgr"] = c
        for key in ["一级分类", "一级", "投资类型", "type1"]:
            if key.lower() in h.lower() and "二级" not in h:
                col_map["type1"] = c
        for key in ["二级分类", "二级", "type2"]:
            if key.lower() in h.lower():
                col_map["type2"] = c

    # 必需列
    if "code" not in col_map:
        # 打印所有列名帮助用户定位
        all_headers = [
            str(ws.cell(row=1, column=c).value or "") for c in range(1, ws.max_column + 1)
        ]
        raise ValueError(
            f"未找到基金代码列。Excel 列名: {all_headers}\n"
            f"请确保包含'基金代码'或'代码'列。"
        )

    funds = []
    for r in range(2, ws.max_row + 1):
        code = str(ws.cell(row=r, column=col_map["code"]).value or "").strip()
        # 清洗代码(去.OF后缀等)
        code = code.replace(".OF", "").replace(".SZ", "").replace(".SH", "")
        if not code or len(code) != 6 or not code.isdigit():
            continue

        name = str(
            ws.cell(row=r, column=col_map.get("name", col_map["code"])).value or ""
        ).strip()
        mgr = str(
            ws.cell(row=r, column=col_map.get("mgr", col_map["code"])).value or ""
        ).strip()
        t1 = str(
            ws.cell(row=r, column=col_map.get("type1", col_map["code"])).value or ""
        ).strip()
        t2 = str(
            ws.cell(row=r, column=col_map.get("type2", col_map["code"])).value or ""
        ).strip()

        funds.append((code, name, mgr, t1, t2))

    return funds


def format_fund_list(funds: list) -> str:
    """格式化基金列表的简要描述"""
    mgrs = set(f[2] for f in funds)
    return f"{len(funds)} 只基金，{len(mgrs)} 家管理人"


def main():
    parser = argparse.ArgumentParser(description="清盘条款分类 — 三级递进流水线")
    parser.add_argument("input", help="基金列表 Excel 文件路径")
    parser.add_argument(
        "--output", default=None, help="输出 Excel 路径(默认: 清盘条款分类.xlsx)"
    )
    parser.add_argument(
        "--work-dir",
        default=None,
        help="工作目录(保存PDF缓存和中间JSON, 默认: 输入文件所在目录)",
    )
    parser.add_argument(
        "--skip-stage3", action="store_true", help="跳过阶段三(CSRC)"
    )
    parser.add_argument(
        "--no-resume",
        action="store_true",
        help="忽略已有结果缓存，从头重新处理全部基金",
    )
    parser.add_argument("--token", default=None, help="Datayes API Token(也可设环境变量DATAYES_TOKEN)")

    args = parser.parse_args()

    # 设置 token
    if args.token:
        os.environ["DATAYES_TOKEN"] = args.token
    if not os.environ.get("DATAYES_TOKEN", "").strip():
        print("[错误] 缺少 DATAYES_TOKEN。请设置环境变量或使用 --token 参数。")
        sys.exit(1)

    try:
        from datayes_api import run_stage as run_stage_1_2
        from gen_excel import generate
        if not args.skip_stage3:
            from csrc_api import run_stage as run_stage_3
        else:
            run_stage_3 = None
    except ModuleNotFoundError as exc:
        print(f"[错误] 缺少 Python 依赖: {exc.name}")
        print("请先安装 liquidation_clause_skill/requirements.txt 中列出的依赖。")
        sys.exit(1)

    # 设置工作目录
    if args.work_dir:
        work_dir = args.work_dir
    else:
        work_dir = os.path.dirname(os.path.abspath(args.input))
    os.makedirs(work_dir, exist_ok=True)

    pdf_dir = os.path.join(work_dir, "contract_pdfs")
    os.makedirs(pdf_dir, exist_ok=True)

    output_path = args.output or os.path.join(work_dir, "清盘条款分类.xlsx")
    cache_path = os.path.join(work_dir, "results_cache.json")

    # ============== 读取输入 ==============
    print(f"\n读取基金列表: {args.input}", flush=True)
    funds = read_fund_list(args.input)
    print(f"  {format_fund_list(funds)}", flush=True)

    # ============== 读取缓存 ==============
    if args.no_resume:
        cached_results, pending_funds = [], list(funds)
        print("  已禁用断点续跑，将处理全部基金", flush=True)
    else:
        cached_results, pending_funds = load_cached_results(cache_path, funds)
        if cached_results:
            print(
                f"  缓存复用: {len(cached_results)} 只，待处理: {len(pending_funds)} 只",
                flush=True,
            )

    checkpoint_by_code = {
        result["code"]: result for result in cached_results
    }
    checkpoint_pending = 0

    def checkpoint_result(result: dict) -> None:
        nonlocal checkpoint_pending
        if not result.get("clauseType"):
            return
        checkpoint_by_code[result["code"]] = result
        checkpoint_pending += 1
        if checkpoint_pending >= CHECKPOINT_INTERVAL:
            save_results_cache(cache_path, list(checkpoint_by_code.values()))
            checkpoint_pending = 0

    # ============== 阶段一: Datayes 基金合同 ==============
    t0 = time.time()
    s1_classified = []
    s1_not_found = []
    if pending_funds:
        print(f"\n{'='*60}", flush=True)
        print("[阶段一] Datayes 基金合同", flush=True)
        print(f"{'='*60}", flush=True)
        s1_classified, s1_not_found = run_stage_1_2(
            pending_funds,
            pdf_dir,
            stage=1,
            on_result=checkpoint_result,
        )
    else:
        print("\n[阶段一] 无需执行(全部命中结果缓存)", flush=True)

    save_results_cache(cache_path, list(checkpoint_by_code.values()))
    if s1_not_found:
        print(f"\n  → 阶段一未分类: {len(s1_not_found)} 只", flush=True)
    elif pending_funds:
        print("\n  → 阶段一全部完成!", flush=True)

    # ============== 阶段二: 替代公告源 ==============
    s2_classified = []
    s2_not_found = []
    if s1_not_found:
        print(f"\n{'='*60}", flush=True)
        print(f"[阶段二] 替代公告源 (处理 {len(s1_not_found)} 只未分类)", flush=True)
        print(f"{'='*60}", flush=True)
        s2_classified, s2_not_found = run_stage_1_2(
            s1_not_found,
            pdf_dir,
            stage=2,
            on_result=checkpoint_result,
        )
        save_results_cache(cache_path, list(checkpoint_by_code.values()))
    else:
        print("\n[阶段二] 无需执行(阶段一或缓存已全覆盖)", flush=True)

    # ============== 阶段三: CSRC 兜底 ==============
    s3_classified = []
    s3_not_found = []
    s3_candidates = filter_unclassified_funds(
        s2_not_found,
        cached_results,
        s1_classified,
        s2_classified,
    )

    if s3_candidates and not args.skip_stage3:
        print(f"\n{'='*60}", flush=True)
        print(
            f"[阶段三] CSRC 证监会兜底 (仅处理 {len(s3_candidates)} 只未分类)",
            flush=True,
        )
        print(f"{'='*60}", flush=True)
        s3_classified, s3_not_found = run_stage_3(
            s3_candidates,
            pdf_dir,
            on_result=checkpoint_result,
        )
        save_results_cache(cache_path, list(checkpoint_by_code.values()))
    elif args.skip_stage3:
        print("\n[阶段三] 已跳过 (--skip-stage3)", flush=True)
        s3_not_found = s3_candidates
    else:
        print("\n[阶段三] 无需执行(没有仍未分类的基金)", flush=True)

    # ============== 合并结果 ==============
    all_results = (
        cached_results
        + s1_classified
        + s2_classified
        + s3_classified
        + [make_unclassified_result(fund) for fund in s3_not_found]
    )
    result_by_code = {result["code"]: result for result in all_results}
    all_results = [
        result_by_code[fund[0]] for fund in funds if fund[0] in result_by_code
    ]
    save_results_cache(cache_path, all_results)

    # ============== 生成 Excel ==============
    elapsed = time.time() - t0
    print(f"\n{'='*60}", flush=True)
    print("生成最终 Excel...", flush=True)
    print(f"{'='*60}", flush=True)

    generate(output_path, all_results)
    print(f"\n总耗时: {elapsed/60:.1f} 分钟", flush=True)

    # ============== 摘要 ==============
    total = len(all_results)
    classified = sum(1 for result in all_results if result.get("clauseType"))
    classified_pct = classified / total * 100 if total else 0
    print(f"\n{'='*60}", flush=True)
    print("最终统计", flush=True)
    print(f"{'='*60}", flush=True)
    print(f"  缓存复用:                  {len(cached_results)} 只", flush=True)
    print(f"  阶段一(Datayes基金合同):   {len(s1_classified)} 只", flush=True)
    print(f"  阶段二(替代公告源):        {len(s2_classified)} 只", flush=True)
    print(f"  阶段三(CSRC证监会):        {len(s3_classified)} 只", flush=True)
    print("  ─────────────────────", flush=True)
    print(f"  可自动分类:                {classified}/{total} ({classified_pct:.1f}%)", flush=True)
    print(f"  未分类:                    {len(s3_not_found)} 只", flush=True)
    if s3_not_found:
        not_found_codes = [fund[0] for fund in s3_not_found]
        print(f"  未分类代码: {', '.join(not_found_codes)}", flush=True)
    print(f"  输出文件: {output_path}", flush=True)
    print(f"{'='*60}", flush=True)


if __name__ == "__main__":
    main()
