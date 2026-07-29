#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
一次性生成最终 Excel(三张表)。
在所有阶段完成后调用。
"""

import os, json
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def generate(output_path: str, results: list):
    """
    从全量结果生成三表 Excel。

    Args:
        output_path: 输出 .xlsx 路径
        results: [{code, name, mgr, type1, type2, clauseType, clauseText, s3Url, source, stage, reason}, ...]
    """
    wb = Workbook()

    # Styles
    hf = Font(name="微软雅黑", bold=True, size=10)
    hfw = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
    df = Font(name="微软雅黑", size=9)
    bf = Font(name="微软雅黑", bold=True, size=10)
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    sfill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    tb = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    ca = Alignment(horizontal="center", vertical="center", wrap_text=True)
    la = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def style_row(ws, r, cols, align=ca):
        for c in range(1, cols + 1):
            ws.cell(row=r, column=c).border = tb
            ws.cell(row=r, column=c).alignment = align

    # ===== Sheet 1: 管理人汇总 =====
    ws1 = wb.active
    ws1.title = "表1-管理人汇总"

    mgrs = defaultdict(lambda: {"total": 0, "type1": 0, "type2": 0, "type3": 0})
    for r in results:
        mgr = r.get("mgr", "未知")
        ct = r.get("clauseType") or ""
        mgrs[mgr]["total"] += 1
        if "类型1" in ct:
            mgrs[mgr]["type1"] += 1
        elif "类型2" in ct:
            mgrs[mgr]["type2"] += 1
        elif "类型3" in ct:
            mgrs[mgr]["type3"] += 1

    h1 = [
        "序号", "管理人", "基金数量",
        "类型1:备案", "占比",
        "类型2:备案+6月大会", "占比",
        "类型3:自动触发终止", "占比",
    ]
    for c, h in enumerate(h1, 1):
        cell = ws1.cell(row=1, column=c, value=h)
        cell.font = hfw
        cell.fill = hfill
        cell.alignment = ca
        cell.border = tb

    tot = {"total": 0, "type1": 0, "type2": 0, "type3": 0}
    row = 2
    for i, (mgr, d) in enumerate(sorted(mgrs.items(), key=lambda x: -x[1]["total"]), 1):
        ws1.cell(row=row, column=1, value=i).font = df
        ws1.cell(row=row, column=2, value=mgr).font = df
        ws1.cell(row=row, column=3, value=d["total"]).font = df
        for j, t in enumerate(["type1", "type2", "type3"]):
            n = d[t]
            pct = n / d["total"] if d["total"] else 0
            ws1.cell(row=row, column=4 + j * 2, value=n).font = df
            ws1.cell(row=row, column=5 + j * 2, value=f"{pct:.1%}").font = df
            tot[t] += n
        tot["total"] += d["total"]
        style_row(ws1, row, 9)
        ws1.cell(row=row, column=2).alignment = la
        row += 1

    # 合计行
    ws1.cell(row=row, column=1, value="合计").font = bf
    ws1.cell(row=row, column=2, value=f"{len(mgrs)}家管理人").font = bf
    ws1.cell(row=row, column=3, value=tot["total"]).font = bf
    for j, t in enumerate(["type1", "type2", "type3"]):
        n = tot[t]
        ws1.cell(row=row, column=4 + j * 2, value=n).font = bf
        ws1.cell(row=row, column=5 + j * 2, value=f"{n / tot['total']:.1%}").font = bf
    for c in range(1, 10):
        ws1.cell(row=row, column=c).fill = sfill
        ws1.cell(row=row, column=c).border = tb
        ws1.cell(row=row, column=c).alignment = ca
    ws1.cell(row=row, column=2).alignment = la

    for c, w in zip(["A", "B", "C", "D", "E", "F", "G", "H", "I"],
                    [6, 24, 10, 14, 14, 16, 16, 14, 14]):
        ws1.column_dimensions[c].width = w

    # ===== Sheet 2: 基金明细 =====
    ws2 = wb.create_sheet("表2-基金明细")
    h2 = [
        "序号", "基金代码", "基金名称", "管理人",
        "一级分类", "二级分类",
        "清盘条款类型", "条款原文", "合同PDF链接",
        "数据来源", "分类阶段", "备注",
    ]
    for c, h in enumerate(h2, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = hfw
        cell.fill = hfill
        cell.alignment = ca
        cell.border = tb

    for i, item in enumerate(sorted(results, key=lambda x: x["code"]), 1):
        vals = [
            i,
            item["code"],
            item.get("name", ""),
            item.get("mgr", ""),
            item.get("type1", ""),
            item.get("type2", ""),
            item.get("clauseType", "未分类"),
            item.get("clauseText", ""),
            item.get("s3Url", ""),
            item.get("source", ""),
            f"阶段{item.get('stage', '?')}",
            item.get("reason", ""),
        ]
        for j, v in enumerate(vals):
            ws2.cell(row=i + 1, column=j + 1, value=v).font = df
        style_row(ws2, i + 1, 12)
        for c in [3, 4, 8, 9, 12]:
            ws2.cell(row=i + 1, column=c).alignment = la

    for c, w in zip(
        ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"],
        [6, 10, 28, 22, 14, 14, 22, 60, 55, 18, 10, 25],
    ):
        ws2.column_dimensions[c].width = w

    # ===== Sheet 3: 统计说明 =====
    ws3 = wb.create_sheet("表3-统计说明")
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 85

    def ct(r):
        v = r.get("clauseType") if isinstance(r, dict) else None
        return str(v) if v else ""

    total = len(results)
    classified = sum(1 for r in results if "类型" in ct(r))
    t1 = sum(1 for r in results if "类型1" in ct(r))
    t2 = sum(1 for r in results if "类型2" in ct(r))
    t3 = sum(1 for r in results if "类型3" in ct(r))
    not_found = total - classified

    s1 = sum(1 for r in results if r.get("stage") == 1 and ct(r))
    s2 = sum(1 for r in results if r.get("stage") == 2 and ct(r))
    s2_contract = sum(
        1
        for r in results
        if r.get("stage") == 2
        and r.get("source") == "基金合同(宽松复判)"
    )
    s2_alternative = s2 - s2_contract
    s3 = sum(1 for r in results if r.get("stage") == 3 and ct(r))
    notes = [
        ("统计范围", f"共 {total} 只基金，{len(mgrs)} 家管理人。"),
        (
            "分类结果",
            f"类型1(备案): {t1} ({t1/total*100:.1f}%) | "
            f"类型2(备案+6月大会): {t2} ({t2/total*100:.1f}%) | "
            f"类型3(自动触发终止): {t3} ({t3/total*100:.1f}%) | "
            f"未分类: {not_found} ({not_found/total*100:.1f}%)",
        ),
        (
            "分阶段统计",
            f"阶段一(Datayes基金合同): {s1}只 | "
            f"阶段二A(同合同宽松复判): {s2_contract}只 | "
            f"阶段二B(替代公告源): {s2_alternative}只 | "
            f"阶段三(CSRC证监会): {s3}只",
        ),
        (
            "数据来源",
            "阶段一&二: 通联数据(Datayes)公告API → S3 PDF下载 → PyMuPDF/pypdf解析。"
            "阶段三: CSRC证监会信息披露平台(advanced_search_report.do) → pypdf解析。",
        ),
        (
            "类型1: 备案",
            "向证监会报告并提交解决方案，但未同时出现10日报告和6个月大会，且不属于直接终止。"
        ),
        (
            "类型2: 备案+6个月大会",
            "连续50或60个工作日触发后，同时明确10个工作日内报告及6个月内召集基金份额持有人大会。"
        ),
        (
            "类型3: 自动触发终止",
            "连续50或60个工作日触发后直接终止/清算且无需大会；阶段三兼容100人门槛旧契约。"
        ),
        (
            "解析引擎",
            "Datayes S3 PDF: PyMuPDF(主力) → 乱码则切pypdf。"
            "CSRC PDF: 仅pypdf(PyMuPDF对CSRC SimSun字体CMap必然乱码)。"
        ),
    ]

    for i, (k, v) in enumerate(notes, 1):
        ws3.cell(row=i, column=1, value=k).font = bf
        ws3.cell(row=i, column=1).fill = sfill
        ws3.cell(row=i, column=1).border = tb
        ws3.cell(row=i, column=1).alignment = Alignment(
            horizontal="center", vertical="top", wrap_text=True
        )
        ws3.cell(row=i, column=2, value=v).font = df
        ws3.cell(row=i, column=2).border = tb
        ws3.cell(row=i, column=2).alignment = Alignment(
            horizontal="left", vertical="top", wrap_text=True
        )
        ws3.row_dimensions[i].height = 45

    wb.save(output_path)

    # 打印统计
    print(f"\n{'='*60}")
    print(f"Excel 已保存: {os.path.basename(output_path)}")
    print(f"  基金总数: {total}")
    print(f"  管理人: {len(mgrs)} 家")
    print(f"  可自动分类: {classified}/{total} ({classified/total*100:.1f}%)")
    print(f"    T1(备案): {t1} ({t1/total*100:.1f}%)")
    print(f"    T2(备案+6月大会): {t2} ({t2/total*100:.1f}%)")
    print(f"    T3(自动触发终止): {t3} ({t3/total*100:.1f}%)")
    print(f"  未分类: {not_found}")
    print(f"  分阶段: S1={s1} S2={s2} S3={s3}")
    print(f"{'='*60}")
